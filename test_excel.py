import openpyxl
import datetime
import random

# Define the Excel file name based on the current timestamp
excel_name = f"data_klimat_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Data'

# Write the headers to the first row of the sheet
ws.cell(row=1, column=1, value='Value')
ws.cell(row=1, column=2, value='Timestamp')

# Insert some random data into the sheet
for i in range(2, 12):
    value = random.uniform(0, 1) # Generate a random value between 0 and 1
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') # Get the current timestamp
    row = (value, timestamp)
    ws.append(row)

# Save the Excel workbook
wb.save(excel_name)

# Close the workbook when done
wb.close()
