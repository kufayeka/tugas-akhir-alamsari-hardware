import openpyxl
import datetime
import random
import time

# Define the Excel file name based on the current timestamp
excel_name = f"data_klimat_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Data'

# Write the headers to the first row of the sheet
ws.cell(row=1, column=1, value='Value')
ws.cell(row=1, column=2, value='Timestamp')

# Insert random data into the sheet every second
while True:
    value = random.uniform(0, 1) # Generate a random value between 0 and 1
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') # Get the current timestamp
    row = (value, timestamp)
    ws.append(row)
    wb.save(excel_name) # Save the Excel workbook
    time.sleep(1) # Wait for 1 second before inserting the next row

# Close the workbook when done
wb.close()
